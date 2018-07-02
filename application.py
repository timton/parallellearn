from flask import Flask, flash, redirect, render_template, request, session, url_for, send_file, after_this_request
from flask_session import Session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import and_, or_
from flask_script import Manager
from flask_migrate import Migrate, MigrateCommand
from passlib.apps import custom_app_context as pwd_context
from tempfile import mkdtemp
from datetime import datetime
from random import shuffle, choice

# for file uploading, manipulation
import os
import urllib.request
import pathlib
from werkzeug.utils import secure_filename

# for excel manipulation
import openpyxl

# for email sending
# https://code.tutsplus.com/tutorials/intro-to-flask-adding-a-contact-page--net-28982
from flask_mail import Message, Mail

from helpers import *

# configure application
# don't track changes (or track them), otherwise error
app = Flask(__name__)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# configure SQLAlchemy database
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ['DATABASE_URL']
db = SQLAlchemy(app)

# mail object
mail = Mail()

# database migration script
migrate = Migrate(app, db)
manager = Manager(app)
manager.add_command('db', MigrateCommand)

# mail configuration
# http://settings.email/corp.mail.ru-email-settings.html
app.config["MAIL_SERVER"] = "smtp.mail.ru"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = 'gunter333@mail.ru'
app.config["MAIL_PASSWORD"] = '1qaz@WSX'

mail.init_app(app)

# configure database models
class Comment(db.Model):
    __tablename__ = 'comments'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False)
    comment = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

    # initially we had sqlite which returns lists of dicts
    # sqlalchemy returns lists of object, so we add a dict method
    # https://stackoverflow.com/questions/35282222/in-python-how-do-i-cast-a-class-object-to-a-dict/35282286
    def asdict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'comment': self.comment,
            'timestamp': self.timestamp
        }

class Correction(db.Model):
    __tablename__ = 'corrections'

    id = db.Column(db.Integer, primary_key=True)
    line_id = db.Column(db.Integer, nullable=False)
    correction = db.Column(db.Text, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    project_id = db.Column(db.Integer, nullable=False)
    version_id = db.Column(db.Integer, nullable=False)
    context_line_id = db.Column(db.Integer, nullable=False)

    def asdict(self):
        return {
            'id': self.id,
            'line_id': self.line_id,
            'correction': self.correction,
            'user_id': self.user_id,
            'project_id': self.project_id,
            'version_id': self.version_id,
            'context_line_id': self.context_line_id
        }

class Line(db.Model):
    __tablename__ = 'lines'

    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, nullable=False)
    version_id = db.Column(db.Integer, nullable=False)
    line_index = db.Column(db.Integer, nullable=False)
    line = db.Column(db.Text, nullable=False)

    def asdict(self):
        return {
            'id': self.id,
            'project_id': self.project_id,
            'version_id': self.version_id,
            'line_index': self.line_index,
            'line': self.line
        }

class Project(db.Model):
    __tablename__ = 'projects'

    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.Text, nullable=False)
    title = db.Column(db.Text, nullable=False)
    author = db.Column(db.Text, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    line_count = db.Column(db.Integer, nullable=False)
    season = db.Column(db.Integer, nullable=True)
    episode = db.Column(db.Integer, nullable=True)
    poster = db.Column(db.Text, nullable=False)
    description = db.Column(db.Text, nullable=False)

    def asdict(self):
        return {
            'id': self.id,
            'type': self.type,
            'title': self.title,
            'author': self.author,
            'year': self.year,
            'user_id': self.user_id,
            'line_count': self.line_count,
            'season': self.season,
            'episode': self.episode,
            'poster': self.poster,
            'description': self.description
        }

class Rating(db.Model):
    __tablename__ = 'ratings'

    id = db.Column(db.Integer, primary_key=True)
    version_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    rating = db.Column(db.Integer, nullable=False)

    def asdict(self):
        return {
            'id': self.id,
            'version_id': self.version_id,
            'user_id': self.user_id,
            'rating': self.rating
        }

class Resumable(db.Model):
    __tablename__ = 'resumables'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False)
    current_line_id = db.Column(db.Integer, nullable=False)
    project_id = db.Column(db.Integer, nullable=False)
    from_version_id = db.Column(db.Integer, nullable=False)
    to_version_id = db.Column(db.Integer, nullable=False)

    def asdict(self):
        return {
            'id': self.id,
            'user_id': self.user_id,
            'current_line_id': self.current_line_id,
            'project_id': self.project_id,
            'from_version_id': self.from_version_id,
            'to_version_id': self.to_version_id
        }

class User(db.Model):
    __tablename__ = 'users'

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.Text, nullable=False, unique=True)
    username = db.Column(db.Text, nullable=False, unique=True)
    hash = db.Column(db.Text, nullable=False)

    def asdict(self):
        return {
            'id': self.id,
            'email': self.email,
            'username': self.username,
            'hash': self.hash
        }

class Version(db.Model):
    __tablename__ = 'versions'

    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    language = db.Column(db.Text, nullable=False)
    source = db.Column(db.Text, nullable=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    rating = db.Column(db.Float, default=0.0)

    def asdict(self):
        return {
            'id': self.id,
            'project_id': self.project_id,
            'user_id': self.user_id,
            'language': self.language,
            'source': self.source,
            'timestamp': self.timestamp,
            'rating': self.rating
        }

# ensure responses aren't cached
if app.config["DEBUG"]:
    @app.after_request
    def after_request(response):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Expires"] = 0
        response.headers["Pragma"] = "no-cache"
        return response

# configure session to use filesystem (instead of signed cookies)
app.config["SESSION_FILE_DIR"] = mkdtemp()
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# uploading files
# http://flask.pocoo.org/docs/0.12/patterns/fileuploads/
UPLOAD_FOLDER = 'static/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# main page route
@app.route("/")
def index():

    # the variable we will pass for rendering
    display = {}

    # get all the supported languages
    all_versions = dict_conversion(Version.query.all())
    possible_languages = []
    for version in all_versions:
        if version["language"] not in possible_languages:
            possible_languages.append(version["language"])

    # rendering variables for quick practice
    display["from_languages"] = sorted(possible_languages)
    display["to_languages"] = sorted(possible_languages, reverse=True)

    # get all the projects
    projects = dict_conversion(Project.query.all())

    # if series, make title include the season & episode
    for project in projects:
        if project["type"].lower() == "series":
            project["title"] += (" (s" + str(project["season"]) + "/e" + str(project["episode"]) +")")

    # get the 5 newest versions
    # https://stackoverflow.com/questions/4186062/sqlalchemy-order-by-descending
    new_versions = dict_conversion(Version.query.order_by(Version.id.desc()).limit(5).all())

    # get all the metadata
    for version in new_versions:
        rows = dict_conversion(Project.query.filter(Project.id == version["project_id"]).all())
        version["type"] = rows[0]["type"]
        version["author"] = rows[0]["author"].split()[-1]
        version["year"] = rows[0]["year"]
        version["title"] = rows[0]["title"]

        if rows[0]["type"].lower() == "series":
            version["title"] += (" (s" + str(rows[0]["season"]) + "/e" + str(rows[0]["episode"]) +")")

        version["timestamp"] = str(version["timestamp"]).split()[0]

    # get the 5 most popular versions
    popular_versions = dict_conversion(Version.query.order_by(Version.rating.desc()).limit(5).all())

    # get all the metadata
    for version in popular_versions:
        rows = dict_conversion(Project.query.filter(Project.id == version["project_id"]).all())
        version["type"] = rows[0]["type"]
        version["author"] = rows[0]["author"].split()[-1]
        version["year"] = rows[0]["year"]
        version["title"] = rows[0]["title"]

        if rows[0]["type"].lower() == "series":
            version["title"] += (" (s" + str(rows[0]["season"]) + "/e" + str(rows[0]["episode"]) +")")

    # get the username, if user logged in
    # https://stackoverflow.com/questions/1602934/check-if-a-given-key-already-exists-in-a-dictionary
    if "user_id" in session:
        rows = dict_conversion(User.query.filter(User.id == session["user_id"]).all())
        username = rows[0]["username"]
    else:
        username = "anon"

    # get all the comments, most recent first
    comments = dict_conversion(Comment.query.order_by(Comment.id.desc()).all())
    for comment in comments:
        if comment["user_id"] == 0:
            comment["username"] = "anon"
        else:
            rows = dict_conversion(User.query.filter(User.id == comment["user_id"]).all())
            comment["username"] = rows[0]["username"]

    # prepare vars for rendering
    # we separate the first project, because the carousel requires a specific active element, so can't use a loop
    if len(projects) > 0:
        display["project_1"] = projects[0]

    if len(projects) > 1:
        display["projects"] = projects[1:]

    display["new_versions"] = new_versions
    display["popular_versions"] = popular_versions
    display["username"] = username
    display["comments"] = comments

    # render the home page, passing in the data
    return render_template("index.html", display=display)


# prepare quick practice route
@app.route("/quick_practice", methods=["GET", "POST"])
def quick_practice():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the desired practice languages
        from_language = request.form.get("from_language").lower()
        to_language = request.form.get("to_language").lower()

        # ensure different languages
        if from_language == to_language:
            return apology("Couldn't load the practice interface.",
                           "Must select different practice languages.")

        # get all the projects that support the 'from language'
        # remove duplicates
        from_projects_ids = []
        rows = dict_conversion(Version.query.filter(Version.language == from_language).all())
        for row in rows:
            from_projects_ids.append(row["project_id"])
        from_projects_ids = list(set(from_projects_ids))

        # get all the projects that support both languages
        possible_projects_ids = []
        for id in from_projects_ids:
            rows = dict_conversion(Version.query.filter(and_(Version.project_id == id,
                                                             Version.language == to_language)).all())
            for row in rows:
                possible_projects_ids.append(row["project_id"])
        possible_projects_ids = list(set(possible_projects_ids))

        # make sure combination possible
        if len(possible_projects_ids) == 0:
            x = "Sorry, this combination (from {f} to {t}) is not supported.".format(f=from_language, t=to_language)
            y = "Please try a different language combination. Thank you!"
            return apology(x, y)


        # get all the potential from-versions
        from_versions_ids = []
        for id in possible_projects_ids:
            rows = dict_conversion(Version.query.filter(and_(Version.project_id == id,
                                                              Version.language == from_language)).all())
            for row in rows:
                from_versions_ids.append(row["id"])

        # get all the potential from-lines
        possible_from_lines = []
        for id in from_versions_ids:
            rows = dict_conversion(Line.query.filter(Line.version_id == id).all())
            for row in rows:
                possible_from_lines.append(row)

        # randomize the from-lines
        shuffle(possible_from_lines)

        # make sure max 100 lines per practice session
        from_lines = possible_from_lines[:100]

        # get the corresponding to-lines
        to_lines = []
        for line in from_lines:

            # get all the potential to-versions
            rows = dict_conversion(Version.query.filter(and_(Version.project_id == line["project_id"],
                                                             Version.language == to_language)).all())

            # pick the to-version at random
            to_version = choice(rows)

            # get the corresponding to-line
            rows = dict_conversion(Line.query.filter(and_(Line.version_id == to_version["id"],
                                                          Line.line_index == line["line_index"])).all())
            to_lines.append(rows[0])

        # append the sources
        for line in from_lines:

            # save the first line separately, as it displays raw html, we will hardcode
            if from_lines.index(line) == 0:
                rows = dict_conversion(Project.query.filter(Project.id == line["project_id"]).all())
                line0_source = rows[0]["title"].title()
                continue

            rows = dict_conversion(Project.query.filter(Project.id == line["project_id"]).all())
            line["line"] += ("<span> (" + rows[0]["title"].title() + ")</span>")

        # prepare project to render
        project = {}
        project["line_count"] = len(from_lines)
        project["from_lines"] = from_lines
        project["to_lines"] = to_lines
        project["starting_line"] = 0
        project["line0_source"] = line0_source

        # to standardize quick & project practice
        project["from_version"], project["to_version"] = {}, {}
        project["from_version"]["language"] = from_language
        project["to_version"]["language"] = to_language

        return render_template("quick_practice.html", project=project)

    # else if user reached route via GET (as by clicking a link or via redirect)
    # go to index
    else:
        return index()

# register route
@app.route("/register", methods=["GET", "POST"])
def register():
    """Register user."""

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure email was submitted
        if not request.form.get("email"):
            return apology("must provide email")

        # ensure username was submitted
        elif not request.form.get("username"):
            return apology("must provide username")

        # ensure password was submitted
        elif not request.form.get("password"):
            return apology("must provide password")

        # ensure password was confirmed
        elif not request.form.get("confirm_password"):
            return apology("must confirm password")

        # ensure passwords match
        elif not request.form.get("password") == request.form.get("confirm_password"):
            return apology("passwords do not match")

        # ensure email uniqueness
        rows = dict_conversion(User.query.filter(User.email == request.form.get("email")).all())
        if rows:
            return apology("try another email")

        # ensure username uniqueness
        rows = dict_conversion(User.query.filter(User.username == request.form.get("username")).all())
        if rows:
            return apology("try another username")

        # try to register new user
        try:
            user = User(email=request.form.get("email"),
                        username=request.form.get("username"),
                        hash=pwd_context.hash(request.form.get("password")))
            db.session.add(user)
            db.session.commit()
        except RuntimeError:
            return apology("error while performing registration")

        msg = Message("Parallellearn Message", sender='gunter333@mail.ru', recipients=[request.form.get("email")])
        msg.body = """
		Hi %s! 

		Thank you for registering.

		Regards,
		Parallellearn
		""" % (request.form.get("name"))

        try:
            mail.send(msg)
        except ConnectionRefusedError:
            pass

        # once successfully registered, log user in automatically
        session["user_id"] = user.id

        # redirect user to home page
        return redirect(url_for("index"))

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("register.html")

# logout route
@app.route("/log_out")
def log_out():
    """Log user out."""

    # forget any user_id
    session.clear()

    # redirect user to welcome page
    return redirect(url_for("index"))

#login route
@app.route("/log_in", methods=["GET", "POST"])
def log_in():
    """Log user in."""

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure email/username was submitted
        if not request.form.get("email_or_username"):
            return apology("must provide email or username")
        else:
            eoru = request.form.get("email_or_username")

        # ensure password was submitted
        if not request.form.get("password"):
            return apology("must provide password")
        else:
            pwd = request.form.get("password")

        # query database for email or username
        rows = dict_conversion(User.query.filter(or_(User.email == eoru, User.username == eoru)).all())

        # ensure email/username exists and password is correct
        if len(rows) != 1 or not pwd_context.verify(pwd, rows[0]["hash"]):
            return apology("invalid email/username and/or password")

        # remember which user has logged in
        session["user_id"] = rows[0]["id"]

        # redirect to home-page or success-page (when editting/saving practice lines)
        if "next" in session:
            next = (session['next'].split("/"))[-1]
            if next == 'edit_line' or next == 'save_progress' or next == 'prepare_to_rate':
                session['next'] = "popup"
                return success("You have successfully logged in.")

        return redirect(url_for('index'))

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:

        # forget any user_id
        session.clear()

        # save the next arg, if possible
        if request.args.get('next'):
            session['next'] = request.args.get('next')

        return render_template("login.html")

# view account details route
@app.route("/view_details")
@login_required
def view_details():

    # fetch the user
    user = dict_conversion(User.query.filter(User.id == session["user_id"]).all())

    return render_template("account_details.html", user=user[0])

# change password route
@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure old password was submitted
        if not request.form.get("old_password"):
            return apology("must provide current password")

        # ensure new password was submitted
        elif not request.form.get("new_password"):
            return apology("must provide new password")

        # ensure new password was confirmed
        elif not request.form.get("confirm_password"):
            return apology("must confirm new password")

        # ensure old password is correct
        rows = dict_conversion(User.query.filter(User.id == session["user_id"]).all())
        if not pwd_context.verify(request.form.get("old_password"), rows[0]["hash"]):
            return apology("current password incorrect")

        # ensure new password is input twice
        elif request.form.get("new_password") != request.form.get("confirm_password"):
            return apology("new password must be input twice")

        # ensure new password is different from the old one
        if pwd_context.verify(request.form.get("new_password"), rows[0]["hash"]):
            return apology("new password must be different from current one")

        # update new password hash
        User.query.filter(User.id == session["user_id"])\
                  .update({"hash": pwd_context.hash(request.form.get("new_password"))})
        db.session.commit()         

        # redirect user to home page
        return redirect(url_for("index"))

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("change_password.html")

# change email route
@app.route("/change_email", methods=["GET", "POST"])
@login_required
def change_email():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure new email was submitted
        if not request.form.get("new_email") or not request.form.get("confirm_new_email"):
            return apology("must provide new email")

        # ensure emails match
        if request.form.get("new_email").lower() != request.form.get("confirm_new_email").lower():
            return apology("emails don't match")

        # ensure new email is different from the current one
        users = dict_conversion(User.query.filter(User.id == session["user_id"]).all())

        if request.form.get("new_email") == users[0]["email"]:
            return apology("new email must be different from current one")

        # ensure email not already registered
        rows = dict_conversion(User.query.filter(User.email == request.form.get("new_email")).all())

        if len(rows) > 0:
            return apology("try another email")

        # update new email
        User.query.filter(User.id == session["user_id"])\
                  .update({"email": request.form.get("new_email")})
        db.session.commit()

        # redirect user to home page
        return redirect(url_for("view_details"))

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("change_email.html")

# change username route
@app.route("/change_username", methods=["GET", "POST"])
@login_required
def change_username():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure new username was submitted
        if not request.form.get("new_username") or not request.form.get("confirm_new_username"):
            return apology("must provide new username")

        # ensure usernames match
        if request.form.get("new_username").lower() != request.form.get("confirm_new_username").lower():
            return apology("usernames don't match")

        # ensure new username is different from the current one
        user = dict_conversion(User.query.filter(User.id == session["user_id"]).all())

        if request.form.get("new_username") == user[0]["username"]:
            return apology("new username must be different from current one")

        # ensure username not already registered
        row = dict_conversion(User.query.filter(User.username == request.form.get("new_username")).all())

        if row:
            return apology("try another username")

        # update new username
        User.query.filter(User.id == session["user_id"]).update({"username": request.form.get("new_username")})
        db.session.commit()

        # redirect user to home page
        return redirect(url_for("view_details"))

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("change_username.html")

# upload new project - we start with the type
@app.route("/new_project_type", methods=["GET", "POST"])
@login_required
def new_project_type():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the new project type, store it in the new project object
        session["new_project"]["type"] = request.form.get("new_project_type").lower()

        return render_template("new_project_metadata.html")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:

        # create a new project dictionary
        session["new_project"] = {}

        # start the uploading process with the type template
        return render_template("new_project_type.html")


# upload new project - project metadata
@app.route("/new_project_metadata", methods=["GET", "POST"])
@login_required
def new_project_metadata():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the new project title, store it in the new project object
        # ensure title of the project was submitted
        if not request.form.get("new_project_title"):
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please provide a title.")
        else:
            session["new_project"]["title"] = request.form.get("new_project_title").lower()

        # for series, save the season and episode
        if session["new_project"]["type"].lower() == "series":
            if not request.form.get("new_project_season") or not request.form.get("new_project_episode"):
                session.pop('new_project', None)
                return apology("Couldn't upload this project.",
                               "Please provide a season and/or an episode number.")
            else:
                session["new_project"]["season"] = request.form.get("new_project_season")
                session["new_project"]["episode"] = request.form.get("new_project_episode")

        # get the new project author, store it in the new project object
        if not request.form.get("new_project_author"):
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please provide an author.")
        else:
            session["new_project"]["author"] = request.form.get("new_project_author").lower()

        # get the new project type, store it in the new project object
        if not request.form.get("new_project_year"):
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please provide a year.")
        else:
            session["new_project"]["year"] = request.form.get("new_project_year")

        # ENSURE FILE SELECTED FOR UPLOAD
        # http://flask.pocoo.org/docs/0.12/patterns/fileuploads/
        # check if the post request has the file part
        if 'file' not in request.files:
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please select a file to upload.")

        # if user does not select file, browser also
        # submit a empty part without filename
        file = request.files['file']
        if not file or file.filename == '':
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please select a file to upload.")

        # ensure file name (extension) allowed
        if forbidden_file(file.filename):
            session.pop('new_project', None)
            return apology("Couldn't upload this project.",
                           "Allowed extensions: xls/xlsx/xlsm/xltx/xltm.")

        # make sure project doesn't exist already
        if session["new_project"]["type"].lower() == "series":
            rows = dict_conversion(Project.query.filter(and_(Project.type == session["new_project"]["type"].lower(),
                                                             Project.title == session["new_project"]["title"].lower(),
                                                             Project.author == session["new_project"]["author"].lower(),
                                                             Project.year == session["new_project"]["year"],
                                                             Project.season == session["new_project"]["season"],
                                                             Project.episode == session["new_project"]["episode"])).all())
        else:
            rows = dict_conversion(Project.query.filter(and_(Project.type == session["new_project"]["type"].lower(),
                                                             Project.title == session["new_project"]["title"].lower(),
                                                             Project.author == session["new_project"]["author"].lower(),
                                                             Project.year == session["new_project"]["year"])).all())

        if len(rows) > 0:
            session.pop('new_project', None)
            return apology("This project already exists.", "Try simply uploading versions for it.")

        # ensure 2 <-> 5 language versions
        workbook = openpyxl.load_workbook(file)
        session["new_project"]["number_of_versions"] = len(workbook.worksheets)

        if session["new_project"]["number_of_versions"] > 5:
            session.pop('new_project', None)
            return apology("Couldn't upload this project.",
                           "Don't overwork - neither yourself, nor the server (maximum 5 language versions allowed per file).")
        elif session["new_project"]["number_of_versions"] < 2:
            session.pop('new_project', None)
            return apology("Couldn't upload this project.",
                           "A new project must have at least two language versions.")

        # make sure all versions have the same number of lines
        session["new_project"]["line_count"] = workbook.worksheets[0].max_row
        for worksheet in workbook:
            if worksheet.max_row != session["new_project"]["line_count"]:
                session.pop('new_project', None)
                return apology("Couldn't upload this project.",
                               "All language versions have to have the same number of lines.")

        # save all the lines
        # https://stackoverflow.com/questions/13377793/is-it-possible-to-get-an-excel-documents-row-count-without-loading-the-entire-d
        session["new_project"]["lines"] = []
        for worksheet in workbook:
            l = []
            for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=session["new_project"]["line_count"]):
                for cell in row:
                    value = str(cell.value)
                    l.append(value)
            session["new_project"]["lines"].append(l)

        return render_template("new_project_versions.html")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return new_project_type()


# upload new project - project versions
@app.route("/new_project_versions", methods=["GET", "POST"])
@login_required
def new_project_versions():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # save language versions and potential sources
        session["new_project"]["versions"] = []
        session["new_project"]["sources"] = []
        for i in range(session["new_project"]["number_of_versions"]):
            session["new_project"]["versions"].append(request.form.get(str(i)).lower())
            if request.form.get("source" + str(i)):
                session["new_project"]["sources"].append(request.form.get("source" + str(i)))
            else:
                session["new_project"]["sources"].append(None)


        # ensure at least two different language versions
        if len(set(session["new_project"]["versions"])) == 1:
            session.pop('new_project', None)
            return apology("Couldn't upload this project.",
                           "New project must have at least two different language versions.")

        return render_template("new_project_formatting.html")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return new_project_type()

# upload new project - project poster and description
@app.route("/new_project_formatting", methods=["GET", "POST"])
@login_required
def new_project_formatting():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

		# get the new project poster link, store it in the new project object
        # ensure poster link of the project was submitted
        if not request.form.get("new_project_poster"):
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please provide a poster image URL.")
        else:
            session["new_project"]["poster"] = request.form.get("new_project_poster")

        # get the new project description, store it in the new project object
        # ensure description of the project was submitted
        if not request.form.get("new_project_description"):
            session.pop('new_project', None)
            return apology("Couldn't upload this project.", "Please provide a description.")
        else:
            session["new_project"]["description"] = request.form.get("new_project_description")

        # make sure description not too long
        if len(session["new_project"]["description"]) > 1000:
            session.pop('new_project', None)
            return apology("Let's not go overboard.", "Description too long (maximum 1000 characters).")

        # upload new project metadata
        try:
            if session["new_project"]["type"] != "series":
                project = Project(type=session["new_project"]["type"], title=session["new_project"]["title"],
                                  author=session["new_project"]["author"], year=session["new_project"]["year"],
                                  user_id=session["user_id"], line_count=session["new_project"]["line_count"],
                                  poster=session["new_project"]["poster"], description=session["new_project"]["description"])
            else:
                project = Project(type=session["new_project"]["type"], title=session["new_project"]["title"],
                                  author=session["new_project"]["author"], year=session["new_project"]["year"],
                                  user_id=session["user_id"], line_count=session["new_project"]["line_count"],
                                  season=session["new_project"]["season"], episode=session["new_project"]["episode"],
                                  poster=session["new_project"]["poster"], description=session["new_project"]["description"])
            db.session.add(project)
            db.session.commit()

        except RuntimeError:
            session.pop('new_project', None)
            return apology("Couldn't save this project in the database.")

        # get the id of the new project
        session["new_project"]["project_id"]=project.id

        # upload new project language versions
        for i in range(session["new_project"]["number_of_versions"]):
            try:
                version = Version(project_id=session["new_project"]["project_id"],
                                  user_id=session["user_id"],
                                  language=session["new_project"]["versions"][i],                                
                                  source=session["new_project"]["sources"][i])
                db.session.add(version)
                db.session.commit()
            except RuntimeError:
                Project.query.filter(Project.id == session["new_project"]["project_id"]).delete()
                db.session.commit()
                Version.query.filter(Version.project_id == session["new_project"]["project_id"]).delete()
                db.session.commit()
                session.pop('new_project', None)
                return apology("Couldn't save the language versions in the database.")

        # get the IDs of the new versions
        # https://stackoverflow.com/questions/10897339/python-fetch-first-10-results-from-a-list
        # https://stackoverflow.com/questions/3940128/how-can-i-reverse-a-list-in-python
        version_ids = []
        rows = dict_conversion(Version.query.filter(Version.project_id == session["new_project"]["project_id"])
                                            .order_by(Version.id.desc()).all())

        for row in rows:
            version_ids.append(row["id"])
        version_ids = version_ids[:session["new_project"]["number_of_versions"]]
        version_ids = list(reversed(version_ids))

        # upload the lines for each version
        # https://stackoverflow.com/questions/522563/accessing-the-index-in-python-for-loops
        for i in range(session["new_project"]["number_of_versions"]):
            for index, line in enumerate(session["new_project"]["lines"][i]):
                try:
                    line = Line(project_id=session["new_project"]["project_id"],
                                version_id=version_ids[i], line_index=index, line=line)
                    db.session.add(line)
                    db.session.commit()
                except RuntimeError:
                    for id in version_ids:
                        Line.query.filter(Line.version_id == id).delete()
                        db.session.commit()
                    Project.query.filter(Project.id == session["new_project"]["project_id"]).delete()
                    db.session.commit()
                    Version.query.filter(Version.project_id == session["new_project"]["project_id"]).delete()
                    db.session.commit()
                    session.pop('new_project', None)
                    return apology("Couldn't save the lines in the database.")

        # if successful, pop the session variable and inform
        # personalize the message just to be kewl
        if session["new_project"]["type"].lower() == "series":
                session["new_project"]["title"] += (" (s" + str(session["new_project"]["season"]) + \
                                                    "/(e" + str(session["new_project"]["episode"]) +")")
        s = "\"" + session["new_project"]["title"].title() + "\"" + " by " + session["new_project"]["author"].title() + \
            " has been successfully uploaded."
        session.pop('new_project', None)

        # so that we don't get any close window message
        session['next'] = ""
        return success(s)

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return new_project_type()

@app.route("/upload_existing", methods=["GET", "POST"])
@login_required
def upload_existing():

    # query for all the existing projects
    existing_projects = dict_conversion(Project.query.all())

    # modify title for series
    # and construct project name to be displayed
    for project in existing_projects:
        if project["type"] == "series":
            project["title"] = (project["title"] + " - s" + str(project["season"]) + \
                               "e" + str(project["episode"]))
        project["name"] = "[" + project["type"] + "] " + project["title"] + " - " + \
                          project["author"] + " (" + str(project["year"]) + ")"




    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure project submitted and ...
        if not request.form.get("project"):
            return apology("Must choose project to update.")
        else:
            project_name = request.form.get("project").lower()

        # ... ensure project exists
        project_id = -1
        for project in existing_projects:
            if project["name"].lower() == project_name:
                project_id = project["id"]
                break
        if project_id == -1:
            return apology("Must choose among existing projects (better select from the dropdown menu).")

        # get project metadata
        rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
        session["existing_project"] = rows[0]

        # ENSURE FILE SELECTED FOR UPLOAD
        # http://flask.pocoo.org/docs/0.12/patterns/fileuploads/
        # check if the post request has the file part
        if 'file' not in request.files:
            session.pop('existing_project', None)
            return apology("Couldn't update this project.", "Please select a file to upload.")

        # if user does not select file, browser also
        # submit a empty part without filename
        file = request.files['file']
        if not file or file.filename == '':
            session.pop('existing_project', None)
            return apology("Couldn't update this project.", "Please select a file to upload.")

        # ensure file name (extension) allowed
        if forbidden_file(file.filename):
            session.pop('existing_project', None)
            return apology("Couldn't update this project.",
                           "Allowed extensions: xls/xlsx/xlsm/xltx/xltm.")

        # ensure 1 <-> 5 language versions
        workbook = openpyxl.load_workbook(file)
        session["existing_project"]["number_of_versions"] = len(workbook.worksheets)

        if session["existing_project"]["number_of_versions"] > 5:
            session.pop('existing_project', None)
            return apology("Couldn't update the project with your file.",
                           "Don't overwork - neither yourself, nor the server (maximum 5 language versions allowed per file).")
        elif session["existing_project"]["number_of_versions"] < 1:
            session.pop('existing_project', None)
            return apology("Couldn't update the project with your file.",
                           "You must contribute with at least one language version.")

        # make sure all the new versions have the same number of lines as the original project
        for worksheet in workbook:
            if worksheet.max_row != session["existing_project"]["line_count"]:
                line_count = session["existing_project"]["line_count"]
                session.pop('existing_project', None)
                return apology("Couldn't update the project with your file.",
                               "New versions have to have the same number of lines as the original project (%s)." % line_count)

        # save all the lines
        # https://stackoverflow.com/questions/13377793/is-it-possible-to-get-an-excel-documents-row-count-without-loading-the-entire-d
        session["existing_project"]["lines"] = []
        for worksheet in workbook:
            l = []
            for row in worksheet.iter_rows(min_row=1, max_col=1, max_row=session["existing_project"]["line_count"]):
                for cell in row:
                    value = str(cell.value)
                    l.append(value)
            session["existing_project"]["lines"].append(l)

        return render_template("existing_project_versions.html")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:

        # return the list of projects for the user to choose from
        return render_template("upload_existing.html", existing_projects=existing_projects)

# add to existing project - project versions
@app.route("/existing_project_versions", methods=["GET", "POST"])
@login_required
def existing_project_versions():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # save language versions and potential sources
        session["existing_project"]["versions"] = []
        session["existing_project"]["sources"] = []
        for i in range(session["existing_project"]["number_of_versions"]):
            session["existing_project"]["versions"].append(request.form.get(str(i)).lower())
            if request.form.get("source" + str(i)):
                session["existing_project"]["sources"].append(request.form.get("source" + str(i)))
            else:
                session["existing_project"]["sources"].append(None)

        # upload new language versions
        # prepare to delete, if something goes wrong
        vids = []
        for i in range(session["existing_project"]["number_of_versions"]):
            try:
                version = Version(project_id=session["existing_project"]["id"],
                                  user_id=session["user_id"], language=session["existing_project"]["versions"][i],
                                  source=session["existing_project"]["sources"][i])
                db.session.add(version)
                db.session.commit()
                vids.append(version.id)
            except RuntimeError:
                for vid in vids:
                    Version.query.filter(Version.id == vid).delete()
                    db.session.commit()
                session.pop('existing_project', None)
                return apology("Couldn't save the new language versions in the database.")

        # get the IDs of the new versions
        # https://stackoverflow.com/questions/10897339/python-fetch-first-10-results-from-a-list
        # https://stackoverflow.com/questions/3940128/how-can-i-reverse-a-list-in-python
        version_ids = []
        rows = dict_conversion(Version.query.filter(Version.project_id == session["existing_project"]["id"])
                                            .order_by(Version.id.desc()).all())

        for row in rows:
            version_ids.append(row["id"])
        version_ids = version_ids[:session["existing_project"]["number_of_versions"]]
        version_ids = list(reversed(version_ids))

        # upload the lines for each version
        # https://stackoverflow.com/questions/522563/accessing-the-index-in-python-for-loops
        for i in range(session["existing_project"]["number_of_versions"]):
            for index, line in enumerate(session["existing_project"]["lines"][i]):
                try:
                    line = Line(project_id=session["existing_project"]["id"],
                                version_id=version_ids[i], line_index=index, line=line)
                    db.session.add(line)
                    db.session.commit()
                except RuntimeError:
                    for id in version_ids:
                        Line.query.filter(Line.version_id == id).delete()
                        db.session.commit()
                    Project.query.filter(Project.id == session["existing_project"]["project_id"]).delete()
                    db.session.commit()
                    Version.query.filter(Version.project_id == session["existing_project"]["project_id"]).delete()
                    db.session.commit()
                    session.pop('existing_project', None)
                    return apology("Couldn't save the new lines in the database.")

        # if successful, pop the session variable and inform
        # personalize the message just to be kewl
        if session["existing_project"]["type"].lower() == "series":
                session["existing_project"]["title"] += (" (s" + str(session["existing_project"]["season"]) + \
                                                        "/(e" + str(session["existing_project"]["episode"]) +")")
        s = "\"" + session["existing_project"]["title"].title() + "\"" + " by " + \
            session["existing_project"]["author"].title() + " has been successfully updated."
        session.pop('existing_project', None)

        # so that we don't get any close window message
        session['next'] = ""
        return success(s)

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return upload_existing()


# route with parameters
# https://stackoverflow.com/questions/14032066/can-flask-have-optional-url-parameters
# display project info
@app.route("/view_project/<id>")
def view_project(id):

    # get project metadata
    rows = dict_conversion(Project.query.filter(Project.id == id).all())
    project = rows[0]

    # if series, make title include the season & episode
    if project["type"].lower() == "series":
        project["title"] += (" (s" + str(project["season"]) + "/e" + str(project["episode"]) +")")

    # get project languages and sources
    rows = dict_conversion(Version.query.filter(Version.project_id == id).all())
    language_list = []
    sources = []
    for row in rows:
        if row["language"] not in language_list:
            language_list.append(row["language"])
        credit = {}
        if row["source"] != None:
            users = dict_conversion(User.query.filter(User.id == row["user_id"]).all())
            credit["user"] = users[0]["username"]
            credit["source"] = row["source"]
            credit["language"] = row["language"]
            sources.append(credit)

    languages = ""
    for language in language_list:
        languages += (language + ", ")
    languages = languages[:-2]
    project["languages"] = languages
    project["sources"] = sources

    return render_template("view_project.html", project=project)


# view account history route
@app.route("/view_history")
@login_required
def view_history():

    # get user's started projects
    started_projects = dict_conversion(Resumable.query.filter(Resumable.user_id == session["user_id"]).all())

    # get all the metadata
    if len(started_projects) > 0:
        for project in started_projects:

            # add the resume id
            project["resume_id"] = project["id"]

            # add "from_version" metadata
            rows = dict_conversion(Version.query.filter(Version.id == project["from_version_id"]).all())
            project["from_language"] = rows[0]["language"]

            # add "to_version" metadata
            rows = dict_conversion(Version.query.filter(Version.id == project["to_version_id"]).all())
            project["to_language"] = rows[0]["language"]

            # add project metadata
            rows = dict_conversion(Project.query.filter(Project.id == project["project_id"]).all())
            project["type"] = rows[0]["type"]
            project["title"] = rows[0]["title"]
            project["author"] = rows[0]["author"]
            project["year"] = rows[0]["year"]
            project["id"] = rows[0]["id"]
            project["current_line_id"] = int(project["current_line_id"] / rows[0]["line_count"] * 100) + 1

            # if series, make title include the season & episode
            if project["type"].lower() == "series":
                project["title"] += (" (s" + str(rows[0]["season"]) + "/e" + str(rows[0]["episode"]) +")")


    # get user's uploaded projects
    uploaded_projects = dict_conversion(Project.query.filter(Project.user_id == session["user_id"]).all())

    # get only the user's versions for his projects
    if len(uploaded_projects) > 0:
        for project in uploaded_projects:

            # if series, make title include the season & episode
            if project["type"].lower() == "series":
                project["title"] += (" (s" + str(project["season"]) + "/e" + str(project["episode"]) +")")

            rows = dict_conversion(Version.query.filter(and_(Version.project_id == project["id"],
                                                             Version.user_id == session["user_id"])).all())

            languages_list = []
            for row in rows:
                if row["language"] not in languages_list:
                    languages_list.append(row["language"])
            languages_list = sorted(languages_list)

            languages_string = ""
            for language in languages_list:
                languages_string += (language + ", ")

            languages_string = languages_string[:-2]
            project["languages"] = languages_string


    # get user's uploaded versions for other users' projects
    user_projects = []
    for project in uploaded_projects:
        user_projects.append(project["id"])

    uploaded_versions = dict_conversion(Version.query.filter(Version.user_id == session["user_id"]).all())

    user_versions = []
    for version in uploaded_versions:
        if version["project_id"] not in user_projects:
            if version["project_id"] not in user_versions:
                user_versions.append(version)

    not_user_project_ids = []
    for version in user_versions:
        if version["project_id"] not in not_user_project_ids:
            not_user_project_ids.append(version["project_id"])

    # get projects' metadata
    not_user_projects = []
    if len(not_user_project_ids) > 0:

        for id in not_user_project_ids:
            project = {}
            project["id"] = id
            rows = dict_conversion(Project.query.filter(Project.id == id).all())
            project["type"] = rows[0]["type"]
            project["title"] = rows[0]["title"]
            project["author"] = rows[0]["author"]
            project["year"] = rows[0]["year"]


            # if series, make title include the season & episode
            if project["type"].lower() == "series":
                project["title"] += (" (s" + str(rows[0]["season"]) + "/e" + str(rows[0]["episode"]) +")")

            # add user's language versions
            rows = dict_conversion(Version.query.filter(and_(Version.project_id == id,
                                                             Version.user_id == session["user_id"])).all())

            languages_list = []
            for row in rows:
                if row["language"] not in languages_list:
                    languages_list.append(row["language"])
            languages_list = sorted(languages_list)

            languages_string = ""
            for language in languages_list:
                languages_string += (language + ", ")

            languages_string = languages_string[:-2]
            project["languages"] = languages_string

            not_user_projects.append(project)

    # render the account history page, passing in the data
    return render_template("account_history.html", started_projects=started_projects,
                           uploaded_projects=uploaded_projects, not_user_projects=not_user_projects)


# prepare to delete
@app.route("/prepare_deletion", methods=["GET", "POST"])
@login_required
def prepare_deletion():

    if request.method == "POST":

        # get the id of the project to delete
        project_id = request.form.get('delete_project')

        # get the project metadata
        rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
        project = rows[0]

        # if series, make title include the season & episode
        if project["type"].lower() == "series":
            project["title"] += (" (s" + str(rows[0]["season"]) + "/e" + str(rows[0]["episode"]) +")")

        project["versions_to_delete"] = dict_conversion(Version.query.filter(and_(Version.project_id == project_id,
                                                                                  Version.user_id == session["user_id"]))
                                                                     .order_by(Version.language.asc()).all())

        return render_template("delete.html", project=project)

    else:
        return view_history()


# delete uploaded project route
@app.route("/delete", methods=["GET", "POST"])
@login_required
def delete():

    # if route reached via POST
    # i.e. one of the delete buttons pressed
    if request.method == "POST":

        # if trying to delete the project entirely
        if request.form.get('delete_project'):

            # get the project id
            project_id = request.form.get('delete_project')

            # bit redundant, but you can only delete your own projects
            rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
            if rows[0]["user_id"] != session["user_id"]:
                return apology("Couldn't delete this project.", "You can only delete your own projects.")


            # if other users contributed to this project, the user can't delete the project itself
            rows = dict_conversion(Version.query.filter(Version.project_id == project_id).all())
            for row in rows:
                if row["user_id"] != session["user_id"]:
                    return apology("Can't delete this project. Other users have contributed to it.",
                                   "Try deleting versions separately instead.")

            # can't delete the project if used by somebody else
            rows = dict_conversion(Resumable.query.filter(Resumable.project_id == project_id).all())
            for row in rows:
                if row["user_id"] != session["user_id"]:
                    return apology("Can't delete this project. Somebody is using it for practice.",
                                   "Try deleting versions separately instead.")

            # project can be deleted, so
            # deleting any corrections first
            try:
                Correction.query.filter(Correction.project_id == project_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't delete this project.", "Please try again.")

            # then the lines
            try:
                Line.query.filter(Line.project_id == project_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't delete this project.", "Please try again.")

            # then the versions
            try:
                Version.query.filter(Version.project_id == project_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't delete this project.", "Please try again.")

            # then the project itself
            try:
                Project.query.filter(Project.id == project_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't delete this project.", "Please try again.")

        # if deleting versions one by one
        elif request.form.get('delete_version'):

            # get the version id
            version_id = request.form.get('delete_version')

            # bit redundant, but you can only delete your own versions
            rows = dict_conversion(Version.query.filter(Version.id == version_id).all())
            if rows[0]["user_id"] != session["user_id"]:
                return apology("Can only delete your own versions.")

            # ''' make sure 2+ different language versions uploaded by the same user are left '''
            # get the project id
            project_id = rows[0]["project_id"]

            # get all the versions of this project
            other_versions = dict_conversion(Version.query.filter(and_(Version.project_id == project_id,
                                                                       Version.id != version_id)).all())

            # if only one version remains
            if len(other_versions) == 1:
                return apology("Couldn't proceed with the deletion.",
                               "At least two different language versions should remain \
                               (for practice purposes).")

            # if 2+ versions remain
            elif len(other_versions) >= 2:

                same_user_different_versions = False
                users_languages = {}

                for version in other_versions:
                    if version["user_id"] not in users_languages:
                        users_languages[version["user_id"]] = version["language"]
                    else:
                        if users_languages[version["user_id"]] != version["language"]:
                            same_user_different_versions = True
                            break

                if not same_user_different_versions:
                    return apology("Couldn't proceed with the deletion.",
                                   "None or at least two different language versions uploaded by the same user should remain.")
            # ''' made sure 2+ different language versions uploaded by the same user are left '''


            # version can be deleted, so
            # deleting any corrections first
            try:
                Correction.query.filter(Correction.version_id == version_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't proceed with the deletion.", "Please try again.")

            # then the lines
            try:
                Line.query.filter(Line.version_id == version_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't proceed with the deletion.", "Please try again.")

            # then the version itself
            try:
                Version.query.filter(Version.id == version_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't proceed with the deletion.", "Please try again.")

            # if no versions left for the project, delete the project itself
            remaining_versions = dict_conversion(Version.query.filter(Version.project_id == project_id)
                                                              .order_by(Version.timestamp.asc()).all())

            if len(remaining_versions) == 0:
                try:
                    Project.query.filter(Project.id == project_id).delete()
                    db.session.commit()
                except RuntimeError:
                    return apology("Deleted the versions, but not the project.", "Please try again.")

            # if no versions of the original user remain
            # change the project user to the next one with the oldest two different versions
            different_user_needed = True
            for version in remaining_versions:
                if version["user_id"] == session["user_id"]:
                    different_user_needed = False
                    break

            if different_user_needed:

                users_languages = {}

                for version in remaining_versions:
                    if version["user_id"] not in users_languages:
                        users_languages[version["user_id"]] = version["language"]
                    else:
                        if users_languages[version["user_id"]] != version["language"]:
                            try:
                                Project.query.filter(Project.id == project_id)\
                                             .update({"user_id": version["user_id"]})
                                db.session.commit()
                            except RuntimeError:
                                return apology("Couldn't change the owner of the project.")
                            break


        # back to account history after deletion
        return view_history()

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to account history to choose project to delete
    else:
        return view_history()


# edit uploaded project route
@app.route("/edit", methods=["GET", "POST"])
@login_required
def edit():

    # if user reached route via POST (as by submitting a form via POST)
    # Different types of POST requests in the same route in Flask
    # https://stackoverflow.com/questions/43333440/different-types-of-post-requests-in-the-same-route-in-flask
    if request.method == "POST":

        # if user prepares to edit project from the account history page
        if request.form.get('want_to_edit_project'):

            project_id = request.form.get('want_to_edit_project')
            rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
            project = rows[0]

            # if series, make title include the season & episode
            if project["type"].lower() == "series":
                project["title"] += (" (s" + str(project["season"]) + "/e" + str(project["episode"]) +")")

            # get user project languages
            rows = dict_conversion(Version.query.filter(and_(Version.project_id == project_id,
                                                             Version.user_id == session["user_id"])).all())

            language_list = []
            for row in rows:
                if row["language"] not in language_list:
                    language_list.append(row["language"])

            languages = ""
            for language in language_list:
                languages += (language + ", ")
            languages = languages[:-2]
            project["languages"] = languages

            return render_template("edit.html", project=project)

        # else if user chooses what he wants to change specifically
        else:

            # if user wants to change type
            if request.form.get('want_to_edit_type'):

                project_id = request.form.get('want_to_edit_type')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                return render_template("edit_type.html", project=project)

            # if user wants to change name
            if request.form.get('want_to_edit_name'):

                project_id = request.form.get('want_to_edit_name')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                return render_template("edit_name.html", project=project)

            # if user wants to change poster
            if request.form.get('want_to_edit_poster'):

                project_id = request.form.get('want_to_edit_poster')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                return render_template("edit_poster.html", project=project)

            # if user wants to change description
            if request.form.get('want_to_edit_description'):

                project_id = request.form.get('want_to_edit_description')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                return render_template("edit_description.html", project=project)

            # if user wants to change one of his project versions
            if request.form.get('want_to_edit_versions'):

                project_id = request.form.get('want_to_edit_versions')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                # get user project versions
                rows = dict_conversion(Version.query.filter(and_(Version.project_id == project_id,
                                                                 Version.user_id == session["user_id"])).all())
                project["versions"] = rows

                return render_template("edit_versions.html", project=project)


            # if user actually changes type
            if request.form.get('change_project_type'):

                # get the project id and the new type
                project_id = request.form.get('change_project_type')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]
                new_type = request.form.get('new_type').lower()

                # new type must differ
                if project["type"] == new_type:
                    return apology("Couldn't edit project.", "Make sure the new type differs from the current one.")

                # if new type series, render the input season/episode template
                if new_type == "series":
                    return render_template("edit_episode.html", project=project)

                # otherwise, make sure project wouldn't exist already in its new form
                rows = dict_conversion(Project.query.filter(and_(Project.type == new_type,
                                                                 Project.title == project["title"],
                                                                 Project.author == project["author"],
                                                                 Project.year == project["year"])).all())

                if len(rows) > 0:
                    return apology("This project already exists.")

                # try to edit
                try:
                    Project.query.filter(Project.id == project_id)\
                                 .update({"type": new_type, "season": None, "episode": None})
                    db.session.commit()
                except RuntimeError:
                    return apology("Couldn't update the project.", "Please try again later.")

                return view_history()

            # if user actually changes type to series
            if request.form.get('change_to_series'):

                # make sure episode submitted
                if not request.form.get('new_episode') or not request.form.get('new_season'):
                    return apology("Couldn't edit project.", "Please input the season and/or episode.")

                # get the project id and the episode/season
                project_id = request.form.get('change_to_series')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]
                new_episode = request.form.get('new_episode')
                new_season = request.form.get('new_season')

                # make sure project wouldn't exist already in its new form
                rows = dict_conversion(Project.query.filter(and_(Project.type == "series",
                                                                 Project.title == project["title"],
                                                                 Project.author == project["author"],
                                                                 Project.year == project["year"],
                                                                 Project.season == new_season,
                                                                 Project.episode == new_episode)).all())

                if len(rows) > 0:
                    return apology("This project already exists.")

                # try to edit
                try:
                    Project.query.filter(Project.id == project_id)\
                                 .update({"type": "series", "season": new_season, "episode": new_episode})
                    db.session.commit()
                except RuntimeError:
                    return apology("Couldn't update the project.", "Please try again later.")

                return view_history()

            # if user actually changes name
            if request.form.get('change_project_name'):

                # get the project id and the episode/season
                project_id = request.form.get('change_project_name')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                # ensure something is being changed
                if not request.form.get('change_title') and not request.form.get('change_author') and \
                not request.form.get('change_year'):
                    if project["type"] == "series":
                        if not request.form.get('change_season') and not request.form.get('change_episode'):
                            return apology("Gotta change something.")
                    else:
                        return apology("Gotta change something.")

                # make sure project to-be-changed isn't changed into an existing project
                if request.form["change_title"]:
                    title = request.form["change_title"].lower()
                else:
                    title = project["title"].lower()

                if request.form["change_author"]:
                    author = request.form["change_author"].lower()
                else:
                    author = project["author"].lower()

                if request.form["change_year"]:
                    year = request.form["change_year"]
                else:
                    year = project["year"]

                if project["type"] == "series":
                    if request.form["change_season"]:
                        season = request.form["change_season"]
                    else:
                        season = project["season"]

                    if request.form["change_episode"]:
                        episode = request.form["change_episode"]
                    else:
                        episode = project["episode"]

                if project["type"] == "series":
                    rows = dict_conversion(Project.query.filter(and_(Project.type == project["type"],
                                                                     Project.title == title,
                                                                     Project.author == author,
                                                                     Project.year == year,
                                                                     Project.season == season,
                                                                     Project.episode == episode)).all())
                else:
                    rows = dict_conversion(Project.query.filter(and_(Project.type == project["type"],
                                                                     Project.title == title,
                                                                     Project.author == author,
                                                                     Project.year == year)).all())

                if len(rows) > 0:
                    return apology("This project already exists.")

                # try to edit the metadata
                try:
                    if request.form["change_title"]:
                        Project.query.filter(Project.id == project_id)\
                                     .update({"title": request.form["change_title"].lower()})
                        db.session.commit()

                    if request.form["change_author"]:
                        Project.query.filter(Project.id == project_id)\
                                     .update({"author": request.form["change_author"].lower()})
                        db.session.commit()

                    if request.form["change_year"]:
                        Project.query.filter(Project.id == project_id)\
                                     .update({"year": request.form["change_year"]})
                        db.session.commit()

                    if project["type"] == "series":
                        if request.form["change_season"]:
                            Project.query.filter(Project.id == project_id)\
                                         .update({"season": request.form["change_season"]})
                            db.session.commit()

                        if request.form["change_episode"]:
                            Project.query.filter(Project.id == project_id)\
                                         .update({"episode": request.form["change_episode"]})
                            db.session.commit()

                except RuntimeError:
                    return apology("Couldn't update the project.", "Please try again later.")

                return view_history()

            # if user actually changes poster
            if request.form.get('change_project_poster'):

                # get the project id, fetch the metadata
                project_id = request.form.get('change_project_poster')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                # ensure new poster link of the project was submitted, save it
                if not request.form.get("new_poster"):
                	return apology("Couldn't edit this project.", "Please provide a new poster image URL.")
                else:
                	new_poster = request.form.get('new_poster')
                
                # try to edit the metadata
                try:
                    Project.query.filter(Project.id == project_id).update({"poster": new_poster})
                    db.session.commit()
                except RuntimeError:
                    return apology("Couldn't update the project.", "Please try again later.")

                return view_history()

            # if user actually changes description
            if request.form.get('change_project_description'):

                # get the project id and the new type
                project_id = request.form.get('change_project_description')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]

                # ensure description of the project was submitted, save it
                if not request.form.get("new_description"):
                    return apology("Couldn't update this project.", "Please provide a new description.")
                else:
                    new_description = request.form.get('new_description')

                # make sure description not too long
                if len(new_description) > 1000:
                    return apology("Didn't edit the project.", "New Description too long (maximum 1000 characters).")

                # try to edit
                try:
                    Project.query.filter(Project.id == project_id).update({"description": new_description})
                    db.session.commit()
                except RuntimeError:
                    return apology("Couldn't update the project.", "Please try again later.")

                return view_history()

            # if user actually changes language version
            if request.form.get('change_project_versions'):

                # get the project id and the new version language
                project_id = request.form.get('change_project_versions')
                rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
                project = rows[0]
                from_version_id = request.form.get('change_from_version')
                to_version = request.form.get('change_to_version').lower()

                # get the potential source
                if request.form.get("source"):
                    source = request.form.get("source")
                else:
                    source = None

                # new version language must differ
                rows = dict_conversion(Version.query.filter(Version.id == from_version_id).all())
                if rows[0]["language"] == to_version:
                    return apology("Didn't update the project.", "New version language must differ from the old one.")

                # make sure we would be left with at least two different language versions for this project
                rows = dict_conversion(Version.query.filter(and_(Version.project_id == project_id,
                                                                 Version.id != from_version_id)).all())
                remaining_languages = []
                for row in rows:
                    remaining_languages.append(row["language"])
                remaining_languages.append(to_version)
                if len(set(remaining_languages)) < 2:
                    return apology("Couldn't update the project.",
                                   "Must have at least two different language versions for every project.")

                # try to edit
                try:
                    Version.query.filter(Version.id == from_version_id)\
                                 .update({"language": to_version, "source": source})
                    db.session.commit()
                except RuntimeError:
                    return apology("Couldn't update the project.", "Please try again later.")

                return view_history()

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to account history to choose project to edit
    else:
        return view_history()


# prepare project for practice route
@app.route("/prepare_practice", methods=["GET", "POST"])
def prepare_practice():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        project_id = request.form["prepare_practice"]
        rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
        project = rows[0]
        project["versions"] = dict_conversion(Version.query.filter(Version.project_id == project_id).all())

        # fetch the user for each version
        for version in project["versions"]:
            rows = dict_conversion(User.query.filter(User.id == version["user_id"]).all())
            version["user"] = rows[0]["username"]

        return render_template("prepare_practice.html", project=project)

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to project browsing page
    else:
        return browse()


# practice route
@app.route("/project_practice", methods=["GET", "POST"])
def project_practice():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # if project started from the project profile page
        if request.form.get('start_practice'):

            # get project metadata
            project_id = request.form.get("start_practice")
            rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
            project = rows[0]

            # ensure translate from & to languages submitted
            if not request.form.get("from_version_id") or not request.form.get("to_version_id"):
                return apology("Couldn't start practice.", "Please choose from & to which language to translate.")
            else:
                from_version_id = request.form.get("from_version_id")
                to_version_id = request.form.get("to_version_id")

            # make sure from & to languages different
            rows = dict_conversion(Version.query.filter(Version.id == from_version_id).all())
            project["from_version"] = rows[0]
            rows = dict_conversion(Version.query.filter(Version.id == to_version_id).all())
            project["to_version"] = rows[0]
            if from_version_id == to_version_id or project["from_version"]["language"] == project["to_version"]["language"]:
                return apology("Couldn't start practice.", "From & to languages must differ.")

            # prepare the starting point
            # if user logged in, check whether project already started
            # https://stackoverflow.com/questions/3845362/python-how-can-i-check-if-the-key-of-an-dictionary-exists
            project["starting_line"] = 0
            if "user_id" in session:
                rows = dict_conversion(Resumable.query.filter(and_(Resumable.user_id == session["user_id"],
                                                                   Resumable.from_version_id == from_version_id,
                                                                   Resumable.to_version_id == to_version_id)).all())

                if len(rows) > 0:
                    project["starting_line"] = rows[0]["current_line_id"]


            # prepare lines and project for rendering
            project["from_lines"] = dict_conversion(Line.query.filter(Line.version_id == from_version_id)
                                                           .order_by(Line.id.asc()).all())
            project["to_lines"] = dict_conversion(Line.query.filter(Line.version_id == to_version_id)
                                                         .order_by(Line.id.asc()).all())


        # else if project practice resumed from the account history page
        elif request.form.get('resume_project'):

            # get resume metadata
            resume_id = request.form.get("resume_project")
            rows = dict_conversion(Resumable.query.filter(Resumable.id == resume_id).all())
            starting_line = rows[0]["current_line_id"]
            from_lines = dict_conversion(Line.query.filter(Line.version_id == rows[0]["from_version_id"])
                                                   .order_by(Line.id.asc()).all())
            to_lines = dict_conversion(Line.query.filter(Line.version_id == rows[0]["to_version_id"])
                                                 .order_by(Line.id.asc()).all())
            project_id = rows[0]["project_id"]

            # get versions metadata
            from_version_id = rows[0]["from_version_id"]
            to_version_id = rows[0]["to_version_id"]
            rows = dict_conversion(Version.query.filter(Version.id == from_version_id).all())
            from_version = rows[0]
            rows = dict_conversion(Version.query.filter(Version.id == to_version_id).all())
            to_version = rows[0]

            # get project metadata
            rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
            project = rows[0]

             # prepare project for rendering
            project["from_lines"] = from_lines
            project["to_lines"] = to_lines
            project["starting_line"] = starting_line
            project["from_version"] = from_version
            project["to_version"] = to_version

        return render_template("project_practice.html", project=project)


    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to project browsing page
    else:
        return browse()

# the about route (the purpose of this website)
@app.route("/about")
def about():
    return render_template("about.html")


# contact info route
####### EMAIL ARE NOT RECEIVED FOR NOW #######
@app.route("/contact", methods=["GET", "POST"])
def contact():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # ensure name was submitted
        if not request.form.get("name"):
            return apology("Must provide name.")

        # ensure username was submitted
        elif not request.form.get("email"):
            return apology("Must provide email.")

        # ensure password was submitted
        elif not request.form.get("message"):
            return apology("Must provide message.")

        msg = Message("Parallellearn Message", sender='gunter333@mail.ru', recipients=['timofei.tonu@gmail.com'])
        msg.body = """
		From: %s <%s>
		%s
		""" % (request.form.get("name"), 
			   request.form.get("email"), 
			   request.form.get("message"))

        try:
            mail.send(msg)
            return success("Thank you for your message! I will try to answer you as soon as possible.")
        except ConnectionRefusedError:
            return apology("Couldn't send your message! Please try again later.")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("contact.html")

# browse route
@app.route("/browse")
def browse():

    # get all the projects, sorted by type
    all_projects = dict_conversion(Project.query.all())
    books = dict_conversion(Project.query.filter(Project.type == "book").all())
    movies = dict_conversion(Project.query.filter(Project.type == "movie").all())
    songs = dict_conversion(Project.query.filter(Project.type == "song").all())

    # if series, make title include the season & episode
    series = []
    for project in all_projects:
        if project["type"].lower() == "series":
            project["title"] += (" (s" + str(project["season"]) + "/e" + str(project["episode"]) +")")
            series.append(project)

    # pass one variable
    projects = {}
    projects["all_projects"] = all_projects
    projects["books"] = books
    projects["series"] = series
    projects["songs"] = songs
    projects["movies"] = movies

    return render_template("browse.html", projects=projects)


# edit version line
@app.route("/edit_line", methods=["GET", "POST"])
@login_required
def edit_line():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # if user prepares to edit line from practice page
        if request.form.get('line_to_edit'):

            # get the line index, and the ids
            line_to_edit = request.form.get("line_to_edit").split(",")
            bad_line_id = line_to_edit[0]
            good_line_id = line_to_edit[1]

            # find out whether the user wants to edit a line of one of hiw own versions
            # if it's hiw own version, he will edit the line directly
            # otherwise, he will submit the correction to the author for consideration
            rows = dict_conversion(Line.query.filter(Line.id == bad_line_id).all())
            bad_version_id = rows[0]["version_id"]
            rows = dict_conversion(Version.query.filter(Version.id == bad_version_id).all())

            if rows[0]["user_id"] == session["user_id"]:
                user_is_author = True
            else:
                user_is_author = False

            # get the lines themselves
            rows = dict_conversion(Line.query.filter(Line.id == good_line_id).all())
            good_line = rows[0]["line"]
            rows = dict_conversion(Line.query.filter(Line.id == bad_line_id).all())
            bad_line = rows[0]["line"]

            # prepare the var to pass
            line_to_edit = {}
            line_to_edit["user_is_author"] = user_is_author
            line_to_edit["good_line"] = good_line
            line_to_edit["good_line_id"] = good_line_id
            line_to_edit["bad_line"] = bad_line
            line_to_edit["bad_line_id"] = bad_line_id

            # render the template with the corresponding variables
            return render_template("edit_line.html", line_to_edit=line_to_edit)

        # else if user actually changes a specific line from the edit line template
        elif request.form.get('edit_line'):

            # make sure new line text provided, save it
            if not request.form.get("edit_line_text"):
                return apology("Nothing has been done.", "Please make sure you submit a correction.")
            else:
                line_text = request.form.get("edit_line_text")

            # get the line index, and the ids
            line_to_edit = request.form.get("edit_line").split(",")
            user_is_author = line_to_edit[0]
            bad_line_id = line_to_edit[1]
            good_line_id = line_to_edit[2]
            rows = dict_conversion(Line.query.filter(Line.id == bad_line_id).all())
            bad_line = rows[0]

            # if user is author, try to edit the line
            if user_is_author == "True":
                try:
                    Line.query.filter(Line.id == bad_line_id).update({"line": line_text})
                    db.session.commit()
                except RuntimeError:
                        return apology("Couldn't edit this line.", "Please try again later.")

                return success("The line has been successfully updated.")

            # else, submit correction to the author
            else:
                try:
                    correction = Correction(line_id=bad_line["id"], correction=line_text, user_id=session["user_id"],
                                            project_id=bad_line["project_id"], version_id=bad_line["version_id"],
                                            context_line_id=good_line_id)
                    db.session.add(correction)
                    db.session.commit()
                except RuntimeError:
                        return apology("Couldn't correct this line.", "Please try again later.")

                return success("The correction has been submitted to the owner of the project.")

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to project browsing page
    else:
        return browse()


# view notifications route
@app.route("/view_notifications")
@login_required
def view_notifications():

    # fetch user's versions
    versions = dict_conversion(Version.query.filter(Version.user_id == session["user_id"]).all())

    # fetch the proposed line corrections for these versions
    corrections = []
    for version in versions:
        rows = dict_conversion(Correction.query.filter(Correction.version_id == version["id"]).all())
        for row in rows:
            corrections.append(row)

    # gather the metadata
    notifications = []
    for correction in corrections:
        notification = {}
        rows = dict_conversion(Project.query.filter(Project.id == correction["project_id"]).all())
        notification["title"] = rows[0]["title"]
        rows = dict_conversion(Line.query.filter(Line.id == correction["context_line_id"]).all())
        notification["context_line"] = rows[0]["line"]
        from_version_id = rows[0]["version_id"]
        rows = dict_conversion(Version.query.filter(Version.id == from_version_id).all())
        notification["from_language"] = rows[0]["language"]
        rows = dict_conversion(User.query.filter(User.id == correction["user_id"]).all())
        notification["user"] = rows[0]["username"]
        notification["suggested_line"] = correction["correction"]
        rows = dict_conversion(Line.query.filter(Line.id == correction["line_id"]).all())
        notification["original_line"] = rows[0]["line"]
        to_version_id = rows[0]["version_id"]
        rows = dict_conversion(Version.query.filter(Version.id == to_version_id).all())
        notification["to_language"] = rows[0]["language"]
        notification["original_line_id"] = correction["line_id"]
        notification["correction_id"] = correction["id"]
        notifications.append(notification)

    return render_template("account_notifications.html", notifications=notifications)


# act upon suggested corrections
@app.route("/action_correction", methods=["GET", "POST"])
@login_required
def action_correction():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # if user accepts the suggested correction
        if request.form.get('implement_correction'):

            correction_id = request.form.get('implement_correction')
            rows = dict_conversion(Correction.query.filter(Correction.id == correction_id).all())
            correction = rows[0]["correction"]
            corrected_line_id = rows[0]["line_id"]

            # change the line
            try:
                Line.query.filter(Line.id == corrected_line_id).update({"line": correction})
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't implement the correction.", "Please try again later.")

            # delete the suggestion
            try:
                Correction.query.filter(Correction.id == correction_id).delete()
                db.session.commit()
            except RuntimeError:
                return apology("Implemented the correction, couldn't delete it though.", "Please try again later.")

        # else if user directly discards the suggested correction
        if request.form.get('discard_correction'):

            # delete the suggestion
            try:
                Correction.query.filter(Correction.id == request.form.get('discard_correction')).delete()
            except RuntimeError:
                return apology("Couldn't delete the correction.", "Please try again later.")

        return view_notifications()

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to project browsing page
    else:
        return browse()


# prepare project for practice route
@app.route("/prepare_download", methods=["GET", "POST"])
def prepare_download():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the project id
        project_id = request.form["prepare_download"]

        # get project
        rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
        project = rows[0]

        # get all its versions
        project["versions"] = dict_conversion(Version.query.filter(Version.project_id == project_id).all())

        # fetch the user for each version
        for version in project["versions"]:
            rows = dict_conversion(User.query.filter(User.id == version["user_id"]).all())
            version["user"] = rows[0]["username"]


        return render_template("prepare_download.html", project=project)

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to project browsing page
    else:
        return browse()


# prepare project for practice route
@app.route("/project_download", methods=["GET", "POST"])
def project_download():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the project id
        project_id = request.form["download_project"]

        # get project
        rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
        project = rows[0]

        # get all its versions
        project["versions"] = dict_conversion(Version.query.filter(Version.project_id == project_id).all())

        # make sure that at least one version was selected for download
        selected_versions = []
        for version in project["versions"]:
            if request.form.get(str(version["id"])):
                selected_versions.append(request.form[str(version["id"])])

        if len(selected_versions) == 0:
            return apology("Couldn't download anything.", "Please select at least one version.")

        # create a new workbook
        wb = openpyxl.Workbook()

        # get the active sheet
        ws = wb.active

        # get the lines for the first version
        lines = dict_conversion(Line.query.filter(Line.version_id == selected_versions[0]).all())

        # fill the first sheet
        # https://stackoverflow.com/questions/31395058/how-to-write-to-a-new-cell-in-python-using-openpyxl
        for line in lines:
            ws.cell(row=(line["line_index"] + 1), column=1).value = line["line"]

        # create and fill additional sheets if more versions
        # https://stackoverflow.com/questions/40385689/add-a-new-sheet-to-a-existing-workbook-in-python
        # https://stackoverflow.com/questions/176918/finding-the-index-of-an-item-given-a-list-containing-it-in-python
        if len(selected_versions) > 1:
            for version in selected_versions:

                # skip the first version, it's done
                if selected_versions.index(version) == 0:
                    continue

                sheet_name = "Sheet" + str(selected_versions.index(version))
                ws = wb.create_sheet(sheet_name)
                lines = dict_conversion(Line.query.filter(Line.version_id == version).all())

                for line in lines:
                    ws.cell(row=(line["line_index"] + 1), column=1).value = line["line"]


        # save the workbook
        try:
            filename = project["title"].title()
            if project["type"].lower() == "series":
                filename += (" (s" + str(project["season"]) + "e" + str(project["episode"]) +")")
            filename += ".xlsx"
            wb.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        except RuntimeError:
            return apology("Couldn't download this project.", "Please try again later.")

        # delete after downloading
        # https://stackoverflow.com/questions/24612366/flask-deleting-uploads-after-they-have-been-downloaded
        @after_this_request
        def remove_file(response):
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            except RuntimeError:
                return apology("Couldn't remove the created file.")
            return response

        # download it
        # https://stackoverflow.com/questions/37937091/download-a-file-from-a-flask-based-python-server
        return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename), as_attachment=True)

    # else if user reached route via GET (as by clicking a link or via redirect)
    # redirect to project browsing page
    else:
        return browse()


# save progress route
@app.route("/save_progress", methods=["GET", "POST"])
@login_required
def save_progress():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the metadata for the progress line
        progress_info = request.form.get("save_progress").split(",")
        line_index = progress_info[0]
        project_id = progress_info[1]
        from_version_id = progress_info[2]
        to_version_id = progress_info[3]

        # if project already started, update the line index
        rows = dict_conversion(Resumable.query.filter(and_(Resumable.user_id == session["user_id"],
                                                           Resumable.project_id == project_id,
                                                           Resumable.from_version_id == from_version_id,
                                                           Resumable.to_version_id == to_version_id)).all())

        if len(rows) > 0:
            try:
            	Resumable.query.filter(Resumable.id == rows[0]["id"]).update({"current_line_id": line_index})
            	db.session.commit()
            except RuntimeError:
                return apology("Couldn't save the progress.", "Please try again later.")

        # otherwise save it as a new started project
        else:
            try:
                resumable = Resumable(user_id=session["user_id"], current_line_id=line_index, project_id=project_id,
                                      from_version_id=from_version_id, to_version_id=to_version_id)
                db.session.add(resumable)
                db.session.commit()
            except RuntimeError:
                    return apology("Couldn't save the progress.", "Please try again later.")

        return success("Progress saved.")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("browse.html")



# discard project route
@app.route("/discard_project", methods=["GET", "POST"])
@login_required
def discard_project():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get discard metadata
        discard_id = request.form.get("discard_project")

        # try deleting the progress
        try:
            Resumable.query.filter(Resumable.id == discard_id).delete()
            db.session.commit()
        except RuntimeError:
            return apology("Couldn't delete the progress.", "Please try again.")

    # view history upon successful discard / GET request
    return view_history()

# faq route
@app.route("/faq")
def faq():
    return render_template("faq.html")


# post message route
@app.route("/comment", methods=["GET", "POST"])
def comment():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the comment
        comment = request.form.get("comment")

        # back to the message board, if empty
        if ''.join(comment.split()) == "":
            return redirect("/#chat-div")

        # check whether user logged in
        if "user_id" in session:
            user_id = session["user_id"]
        else:
            user_id = 0

        # try saving it to the database
        try:
            comment = Comment(user_id=user_id, comment=comment)
            db.session.add(comment)
            db.session.commit()
        except RuntimeError:
            return apology("Couldn't post the message.", "Please try again later.")

    # go back to the message board once posted
    return redirect("/#chat-div")


# check if logged in user has notifications
# more convenient to declare here than in helpers, as would need to move imports
def has_notifications():

    if session:
        if "user_id" in session:

            user_versions = []
            rows = dict_conversion(Version.query.filter(Version.user_id == session["user_id"]).all())

            for row in rows:
                user_versions.append(row["id"])

            notifications = []
            rows = dict_conversion(Correction.query.all())
            for row in rows:
                notifications.append(row["version_id"])

            for version in user_versions:
                if version in notifications:
                    return True
    return False

# https://stackoverflow.com/questions/6036082/call-a-python-function-from-jinja2
app.jinja_env.globals.update(has_notifications=has_notifications)

# prepare to rate a project version
@app.route("/prepare_to_rate", methods=["GET", "POST"])
@login_required
def prepare_to_rate():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the id of the version to rate
        version_id = request.form.get("rating_id")

        # check whether user already rated this version
        rows = dict_conversion(Rating.query.filter(and_(Rating.user_id == session["user_id"],
                                                        Rating.version_id == version_id)).all())

        if len(rows) == 0:
            already_rated = False
            current_rating = None
        else:
            already_rated = True
            current_rating = rows[0]["rating"]

        # get the version metadata
        rows = dict_conversion(Version.query.filter(Version.id == version_id).all())
        language = rows[0]["language"]
        user_id = rows[0]["user_id"]
        project_id = rows[0]["project_id"]
        rows = dict_conversion(User.query.filter(User.id == user_id).all())
        username = rows[0]["username"]
        rows = dict_conversion(Project.query.filter(Project.id == project_id).all())
        version = rows[0]
        if version["type"].lower() == "series":
            version["title"] += (" (s" + str(version["season"]) + "/e" + str(version["episode"]) +")")

        version["id"] = version_id
        version["already_rated"] = already_rated
        version["current_rating"] = current_rating
        version["language"] = language
        version["username"] = username

        return render_template("rate.html", version=version)

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("browse.html")

# rate a project version
@app.route("/rate", methods=["GET", "POST"])
@login_required
def rate():

    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        # get the version id, rating (0-indexed)
        verating = request.form.get("rate")
        verating = verating.split(",")
        version_id = int(verating[0])
        rating = int(verating[1]) + 1

        # check whether user already rated this version
        rows = dict_conversion(Rating.query.filter(and_(Rating.user_id == session["user_id"],
                                                        Rating.version_id == version_id)).all())

        # insert new rating or update existing rating
        if len(rows) == 0:
            try:
                rating = Rating(version_id=version_id, user_id=session["user_id"], rating=rating)
                db.session.add(rating)
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't save rating.", "Please try again later.")
        else:
            try:
                Rating.query.filter(Rating.id == rows[0]["id"]).update({"rating": rating})
                db.session.commit()
            except RuntimeError:
                return apology("Couldn't update rating.", "Please try again later.")

        # calculate the new version rating
        rows = dict_conversion(Rating.query.filter(Rating.version_id == version_id).all())
        new_rating = 0.0
        for row in rows:
            new_rating += row["rating"]
        new_rating = round(new_rating / len(rows), 2)

        # update version rating
        try:
            Version.query.filter(Version.id == version_id).update({"rating": new_rating})
            db.session.commit()
        except RuntimeError:
            return apology("Couldn't update version rating.", "Please try again later.")

        return success("Thank you for rating this version.")

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("browse.html")


if __name__ == '__main__':
	manager.run()

######################
# DONE UP UNTIL HERE #
######################













'''
# prototype
@app.route("/whatever", methods=["GET", "POST"])
@login_required
def whatever():
    # if user reached route via POST (as by submitting a form via POST)
    if request.method == "POST":

        pass

    # else if user reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("upload_new.html")
'''